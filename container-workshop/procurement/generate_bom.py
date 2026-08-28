#!/usr/bin/env python3
"""Container Workshop rev C — procurement BOM workbook generator.

Single source of truth for the RFQ bill of materials issued to prefabrication
builders and suppliers. Emits Container-Workshop_BOM_RevC.xlsx (5 sheets) and
bom_summary.json (group rollup consumed by the RFQ brief generator).

Quantities and dimensions follow the rev C design package (2026-08-24).
Rate and amount columns are deliberately empty: no prices are invented
anywhere in this package — they come from supplier quotes.
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RFQ_NO = "RFQ-CW-2026-001"
ISSUE_DATE = "2026-08-28"
QUOTES_DUE = "2026-09-25"
CONTACT = "aaron@carbonproject.com.au"

# scope codes
CN = "CN"          # China supply — quote required
OPT = "CN-OPT"     # China supply — priced option, excluded from base total
LOCAL = "LOCAL"    # Australian local supply / site works — info only, not in RFQ

# (description, specification & certification, notes, unit, qty, nominal_total_mass_kg, scope)
GROUPS = [
    ("A", "ISO containers & inter-connection", [
        ("40 ft high-cube ISO shipping container, one-trip condition",
         "ISO 668 1AAA, 12.192 x 2.438 x 2.896 m, CSC safety-approval plated, doors and seals intact",
         "The three containers form the building and double as the export freight containers for this kit (shipper-owned containers)",
         "ea", 3, 11700, CN),
        ("Inter-container corner connection plate set, bolted",
         "HDG to AS/NZS 4680; engages ISO corner castings top and bottom",
         "Containers meet at both rear corners of the 12.192 x 12.192 m bay",
         "set", 4, 120, CN),
        ("Container-to-slab hold-down bracket",
         "HDG cleat to suit bottom corner casting, M16 chemical-anchor fixing",
         "4 per container; anchors in group X",
         "ea", 12, 96, CN),
        ("Container survey, prep and rail inspection report",
         "Top side-rail straightness and corner-casting condition recorded per unit",
         "Top side rails carry the storage-deck bearers — condition is structural",
         "ea", 3, None, CN),
        ("Container repaint system, exterior and cut edges",
         "2-pack polyurethane over epoxy primer, marine grade; colour advised at order",
         "Applied after modification group B",
         "lot", 1, None, CN),
    ]),
    ("B", "Container modification & bay structure", [
        ("Side-wall opening, cut and dressed, 2.288 m w x 2.500 m h",
         "Cut lines sealed and primed; corrugation edges dressed",
         "15 openings = 85.8 m2 of stressed skin removed; structural sign-off by principal's engineer before fabrication",
         "ea", 15, None, CN),
        ("Goalpost reinforcement frame at opening (2 posts + head)",
         "150 PFC G300, welded frame, HDG; weld to AS/NZS 1554.1 SP",
         "One frame per opening restores racking stiffness lost with the skin",
         "set", 15, 2100, CN),
        ("Opening perimeter trim / jamb-head flashing set",
         "1.2 BMT folded galv steel, colour-matched",
         None,
         "set", 15, 150, CN),
        ("Head deflection plate and packer set",
         "10 mm plate + slotted packers",
         "Isolates goalpost head from container top-rail deflection",
         "set", 15, 90, CN),
        ("Server compartment fit-out: sealed insulated lining + access door",
         "Stud-framed lining, 50 mm insulation, sealed gasketed door 820 mm",
         "1 of the 15 bays is the sealed server compartment",
         "lot", 1, None, CN),
        ("Server compartment fire-rated lining upgrade",
         "FRL 60/60/60 board system to compartment walls and ceiling",
         "Certifier acceptance path confirmed by principal",
         "lot", 1, None, CN),
        ("Split air-conditioner for server compartment, 2.5 kW, pre-wired",
         "RCM marked; R32; wall-mount head + external condenser bracket",
         None,
         "ea", 1, 45, CN),
        ("Inter-container junction weather seal, full height",
         "EPDM membrane + folded cover flashing",
         "2 junctions at the rear corners",
         "run", 2, None, CN),
        ("Container-top edge kerb angle",
         "75 x 75 x 6 EA G300, HDG, drilled for bearer fixings",
         "Perimeter of the 3.001 m storage deck, approx. 44 m",
         "lot", 1, 300, CN),
        ("Container door retention / latch-open hardware set",
         "Galv hold-open stays + padlockable keepers",
         "Existing container end doors retained",
         "set", 6, 30, CN),
        ("Vent blank-off and roof penetration patch",
         "Seal-welded patches, primed",
         None,
         "lot", 1, None, CN),
        ("In-bay cable duct and lighting channel, per bay",
         "Galv slotted duct 75 x 50 with lid, bay length",
         "Feeds bench GPO rail and bay strip light",
         "ea", 15, 120, CN),
        ("Container, bay and plant signage set",
         "Engraved traffolyte + reflective bay numbers",
         None,
         "lot", 1, None, CN),
    ]),
    ("C", "Primary steel — portal frames", [
        ("Portal column 360 UB 44.7, cut length approx. 6.90 m",
         "AS/NZS 3679.1 grade 300; fabrication AS/NZS 5131 CC2",
         "6 frames x 2 columns; eave height 6.896 m",
         "ea", 12, 3701, CN),
        ("Portal rafter 360 UB 44.7, cut length approx. 8.85 m",
         "AS/NZS 3679.1 grade 300; 10.0 deg pitch; span 17.420 m",
         "6 frames x 2 rafters",
         "ea", 12, 4747, CN),
        ("Eave haunch cutting ex 360 UB 44.7, approx. 1.75 m, with web plate",
         "Profile-cut ex parent section; weld AS/NZS 1554.1 SP",
         "Haunch soffit at 6.082 m governs crane clearance — hold this dimension",
         "ea", 12, 900, CN),
        ("Apex haunch / connection plate pair",
         "20 mm plate grade 350, match-drilled",
         None,
         "set", 6, 240, CN),
        ("Column base plate 350 x 350 x 25 with stiffeners",
         "Grade 250 plate, shop-welded to column",
         None,
         "ea", 12, 660, CN),
        ("Gable / end-wall post 250 UB 25.7, lengths 6.0 - 8.4 m",
         "AS/NZS 3679.1 grade 300",
         "Front and rear gables; supports girts and door framing",
         "ea", 6, 1080, CN),
        ("Roller-door portal: jamb posts + head beam ex 250 UB",
         "Head gives 2.875 m clear above the 4.0 x 4.0 m opening",
         None,
         "set", 1, 400, CN),
        ("Holding-down bolt assembly M24 grade 4.6 HDG, cast-in, with template",
         "AS/NZS 4680 galv; 4 bolts per column + ply setting template",
         "Templates ship first if requested — pits and slab are local works",
         "set", 12, 120, CN),
        ("Structural bolt sets M20/M24 PC 8.8 HDG",
         "AS/NZS 1252 assemblies, batch-certified",
         "All frame, splice and bracket connections",
         "lot", 1, 350, CN),
        ("Purlin, girt, fly-brace and bracket cleats, shop-welded",
         "Grade 250/300 fittings",
         None,
         "lot", 1, 450, CN),
        ("Fabrication allowance: connection plates, stiffeners, shims",
         "Grade 250/300",
         "Balancing allowance within the 20.7 t scheduled steel",
         "lot", 1, 240, CN),
        ("Surface protection, all group C-F steel",
         "Abrasive blast Sa2.5; HDG AS/NZS 4680, or duplex HDG + PU topcoat where noted",
         "Quote as rate per tonne; mass carried on member lines",
         "lot", 1, None, CN),
    ]),
    ("D", "Secondary steel — purlins, girts, canopies", [
        ("Roof purlin Z200-19, punched, with laps",
         "AS 1397 G450 Z350; AS/NZS 4600 design",
         "15 purlin lines front to back; 219.5 m net + laps",
         "m", 242, 1375, CN),
        ("Wall girt Z200-15, punched",
         "AS 1397 G450 Z350",
         "Side walls and both gables",
         "m", 240, 1070, CN),
        ("Eave strut C200-19",
         "AS 1397 G450 Z350",
         "Both eaves, 14.63 m each",
         "m", 29.3, 166, CN),
        ("Ridge purlin pair + ridge ties",
         "Z200-19 + tie straps",
         None,
         "lot", 1, 120, CN),
        ("Roof purlin bridging rows",
         "Proprietary bridging, 3 rows per slope",
         None,
         "lot", 1, 180, CN),
        ("Wall girt bridging",
         "Proprietary bridging",
         None,
         "lot", 1, 120, CN),
        ("Fly braces, rafter to purlin",
         "40 x 40 x 4 EA, HDG",
         None,
         "ea", 48, 110, CN),
        ("Personnel-door frames and head/sill trimmers",
         "Folded galv steel frames for 2 doorsets",
         None,
         "lot", 1, 160, CN),
        ("Canopy frame over roller door, approx. 4.5 x 1.2 m cantilever",
         "SHS 75 x 75 x 4 frame, HDG",
         "Cladding in group I",
         "ea", 1, 220, CN),
        ("Canopy frame over personnel door, approx. 1.5 x 1.0 m",
         "SHS 65 x 65 x 3 frame, HDG",
         None,
         "ea", 1, 60, CN),
        ("Service-reel hood frame, approx. 3.2 x 0.7 m, front wall",
         "SHS 50 x 50 x 3 + folded hood support brackets",
         "Reel bank sits 0.504 m clear of the roller-door opening",
         "ea", 1, 110, CN),
    ]),
    ("E", "Bracing", [
        ("Roof plane rod bracing M20 with turnbuckles",
         "Grade 4.6 rod, HDG; 2 braced bays, both slopes",
         None,
         "lot", 1, 260, CN),
        ("Wall K-bracing SHS 89 x 89 x 5",
         "AS/NZS 1163 C350, HDG; 2 walls x 2 bays",
         None,
         "lot", 1, 480, CN),
        ("Longitudinal strut tubes CHS 114 x 4.5 (apex + mid-slope lines)",
         "AS/NZS 1163 C350, HDG; 3 lines x 14.63 m",
         None,
         "m", 44, 535, CN),
        ("Gable bracing rods and fittings",
         "M16 rod, HDG",
         None,
         "lot", 1, 120, CN),
        ("Bracing cleats and gusset plates",
         "Grade 250 plate",
         None,
         "lot", 1, 90, CN),
        ("Turnbuckles and clevises M20",
         "Rated, galv",
         None,
         "ea", 24, 60, CN),
    ]),
    ("F", "Crane runway", [
        ("Runway beam 310 UB 46.2, two-span continuous lengths approx. 7.55 m",
         "AS/NZS 3679.1 grade 300; design interface AS 1418.18",
         "2 runs x 14.8 m; rail top of steel 5.254 m",
         "ea", 4, 1395, CN),
        ("Runway support corbel bracket with column web stiffening",
         "Shop-welded to portal columns, match-marked",
         "12 brackets, one per column",
         "ea", 12, 540, CN),
        ("Crane rail, 50 x 30 machined flat bar",
         "Grade 350, drilled for clips",
         "2 runs x 14.8 m",
         "m", 29.6, 349, CN),
        ("Rail clips with resilient pads",
         "Proprietary clip system",
         None,
         "lot", 1, 60, CN),
        ("Runway splice plate sets with bolts",
         "Grade 350 plate, PC 8.8 bolts",
         None,
         "set", 6, 90, CN),
        ("Runway end stops with elastomeric buffers",
         "Rated for 2 t crane at full LT speed",
         None,
         "ea", 4, 72, CN),
        ("Runway alignment shim packs and survey record",
         "Stainless shim packs; alignment tolerance per AS 1418.18",
         None,
         "lot", 1, 30, CN),
        ("Festoon / conductor support brackets along runway",
         "Galv brackets at 2 m centres",
         None,
         "lot", 1, 40, CN),
    ]),
    ("G", "Overhead crane, 2000 kg", [
        ("Single-girder EOT crane, SWL 2000 kg, span 17.420 m",
         "AS 1418.1; duty FEM/ISO M5; girder deflection limit 29 mm (design 26 mm)",
         "Hook height 5.154 m; reaches 1.914 m over each container top",
         "ea", 1, 3000, CN),
        ("Electric wire-rope/chain hoist 2 t, dual speed, lift approx. 5.2 m",
         "Duty M5; low-headroom trolley",
         "May be integral to G-01 — state make/model",
         "ea", 1, 250, CN),
        ("End carriages with dual-speed LT drives and soft start",
         "Machined wheels to suit F-03 rail",
         None,
         "set", 1, 300, CN),
        ("Cross-travel and long-travel festoon / energy chain system",
         "Flat cable festoon, galv track",
         None,
         "set", 1, 60, CN),
        ("Radio remote control, hold-to-run, + wired pendant backup",
         "Hold-to-run is a life-safety requirement — no latching motion controls",
         None,
         "set", 1, 10, CN),
        ("Overload limiter, height limit and slack-rope protection",
         "Calibrated, test certificates supplied",
         None,
         "set", 1, 10, CN),
        ("LT travel limit switches and end-of-travel damping",
         None, None,
         "set", 1, 15, CN),
        ("Crane isolator and runway feed",
         "Lockable isolator; supply cable to festoon",
         None,
         "set", 1, 25, CN),
        ("Commissioning spares kit",
         "Contactors, limit switch, brake pads, wheel",
         "125% load test performed in Australia at commissioning",
         "lot", 1, 30, CN),
    ]),
    ("H", "Envelope — PIR panels, flashings, rainwater", [
        ("Roof sandwich panel, 100 mm PIR core, trapezoidal outer skin",
         "0.5/0.4 BMT colour steel faces; FM 4880 approval or AS 1530.4 full test evidence required",
         "Both slopes; slope length 8.845 m; U-value <= 0.25 W/m2K",
         "m2", 266, 3325, CN),
        ("Wall and gable sandwich panel, 100 mm PIR core",
         "As H-01; vertical lay",
         "607 m2 total envelope with H-01",
         "m2", 341, 4265, CN),
        ("Ridge capping, vented, with profile closures",
         "0.55 BMT folded, colour-matched",
         None,
         "m", 14.7, 40, CN),
        ("Barge cappings",
         "0.55 BMT folded",
         "4 x 8.9 m",
         "m", 36, 90, CN),
        ("Corner flashings",
         "0.55 BMT folded",
         None,
         "m", 28, 60, CN),
        ("Base / drip flashing and panel base channel",
         "Galv channel + folded drip",
         "Full perimeter approx. 63.4 m",
         "m", 64, 200, CN),
        ("Eaves gutter, high-capacity, with brackets",
         "0.55 BMT; brackets at 600 crs",
         "Both eaves",
         "m", 29.3, 120, CN),
        ("Downpipes 100 x 75 with astragals",
         "Colour steel",
         "4 off, eave height 6.9 m",
         "m", 28, 60, CN),
        ("Roller and personnel door head/jamb flashing sets",
         "0.55 BMT folded",
         None,
         "set", 2, 40, CN),
        ("Panel fasteners: load-bearing and stitching screws",
         "Class 4 coating, EPDM sealed",
         None,
         "lot", 1, 120, CN),
        ("Butyl tape, gunnable sealant, closed-cell closures",
         "Panel-system approved",
         None,
         "lot", 1, 60, CN),
        ("Rear-gable exhaust fan penetration trims and weather cowls",
         "0.55 BMT + mesh",
         "Fans moved off the roof so nothing penetrates the array",
         "set", 2, 30, CN),
        ("Exhaust fan, 450 mm axial, 2-speed, gable-mounted",
         "RCM marked; controls interface to group U",
         None,
         "ea", 2, 50, CN),
    ]),
    ("I", "Doors", [
        ("Insulated roller shutter 4.0 x 4.0 m, motorised",
         "Wind-rated per AS/NZS 1170.2 site category (TBC); 3-phase drive; manual haul-chain backup",
         "2.875 m clear above the head beam",
         "ea", 1, 900, CN),
        ("Roller door safety package: bottom edge, photocells, battery release",
         "Monitored safety edge; RCM marked controls",
         None,
         "set", 1, 15, CN),
        ("Personnel doorset 920 x 2040, insulated steel leaf",
         "Lever egress hardware — mechanical, no power or credential to exit (life-safety)",
         None,
         "ea", 2, 160, CN),
        ("Door closer and weather seal set",
         None, None,
         "set", 2, 10, CN),
        ("Electric strike, entry side only, egress unaffected",
         "Fail-secure entry; mechanical lever egress always free",
         "Works with group U credential readers",
         "ea", 2, 4, CN),
        ("Roller-door canopy cladding and flashings",
         "PIR / colour steel to match envelope (frame in D-09)",
         None,
         "set", 1, 90, CN),
        ("Personnel-door canopy cladding and flashings",
         "As I-06 (frame in D-10)",
         None,
         "set", 1, 30, CN),
        ("Security hardware: padbolts, hasps, key cylinders keyed alike",
         None, None,
         "lot", 1, 10, CN),
    ]),
    ("J", "Container-top storage deck & racking", [
        ("Deck bearer RHS 100 x 50 x 4, length approx. 2.55 m",
         "AS/NZS 1163 C350, HDG; spans container width onto both top side rails",
         "33 bearers at 1219 mm centres; design UDL 2.5 kPa (subframe design to be verified by principal's engineer)",
         "ea", 33, 715, CN),
        ("Storage-zone grating, F/L 32 x 5, HDG",
         "AS 1657 compliant walking surface; banded panels",
         "Approx. 55.5 m2 of the 89.2 m2 deck",
         "m2", 55.5, 1560, CN),
        ("Walkway grating, 900 mm wide run",
         "As J-02",
         "33.7 m2 around the deck inner edge",
         "m2", 33.7, 945, CN),
        ("Grating hold-down clips and fixings",
         "Galv saddle clips",
         None,
         "lot", 1, 40, CN),
        ("Deck tie-down / lashing points, 1 t rated",
         "Bolt-on, rated and tagged",
         None,
         "ea", 12, 36, CN),
        ("Racking end frames, 1.5 m h x 0.6 m d, HDG",
         "AS 4084; height limited to 1.5 m by crane girder clearance",
         "5 bays, total capacity 7500 kg",
         "ea", 6, 150, CN),
        ("Racking beam pairs, 2.4 m, 750 kg per level",
         "AS 4084, safety locks",
         "2 levels per bay",
         "pr", 10, 180, CN),
        ("Racking mesh deck panels",
         "Galv mesh, 1200 x 600",
         None,
         "ea", 20, 160, CN),
        ("Racking base plates, deck fixings and SWL signage",
         "AS 4084 load signs",
         None,
         "lot", 1, 20, CN),
        ("Stored-item height gauge and 1.903 m envelope marking",
         "Powder-coated gauge bar + line marking",
         "Max craned item height on the deck is 1.903 m",
         "lot", 1, 15, CN),
    ]),
    ("K", "Stair, walkway edge protection", [
        ("Stair, 2 flights x 8 risers, 900 mm wide, grating treads",
         "AS 1657; rise 188 mm, going 216 mm; total rise 3.008 m",
         "Fits the 2.438 m corner void with 128 mm to spare",
         "ea", 1, 260, CN),
        ("Mid landing 900 x 900 with posts",
         "AS 1657; grating top",
         None,
         "ea", 1, 60, CN),
        ("Stair balustrade, both sides, with handrails",
         "AS 1657, HDG",
         None,
         "lot", 1, 90, CN),
        ("Deck edge guardrail: top + mid rail + kickplate",
         "40 NB CHS, HDG, AS 1657; stanchions to bearers",
         "Approx. 40 m of open edge",
         "m", 40, 320, CN),
        ("Self-closing gate at stair head",
         "AS 1657, sprung hinges",
         None,
         "ea", 1, 25, CN),
        ("Maintenance anchor points, roof/array access",
         "AS/NZS 1891 rated, tagged, 4 off",
         None,
         "ea", 4, 20, CN),
        ("Guardrail and stair fixing sets",
         "HDG bolt sets",
         None,
         "lot", 1, 25, CN),
        ("Deck safety signage: 2.5 kPa SWL, stair and gate signs",
         None, None,
         "lot", 1, 5, CN),
    ]),
    ("L", "In-ground vehicle lifts, 4 x 6 t", [
        ("In-ground telescopic lift cassette, 6000 kg, 3-stage, 3.0 m stroke",
         "Stage bores 180/140/100 mm; buckling SF 4.57 at 6 t full extension; AS/NZS 1418.9",
         "Cassette includes steel pit box for casting-in; pits on a 4.0 x 3.5 m grid",
         "ea", 4, 3400, CN),
        ("Flush drive-over cover system per lift",
         "Load-rated covers, anti-slip, flush with slab",
         None,
         "set", 4, 720, CN),
        ("Removable cross beams / pick-up adaptors with storage trolley",
         "Rated 6 t per lift position",
         None,
         "set", 4, 480, CN),
        ("Rack-and-pawl mechanical lock assembly per lift",
         "Engages under load independent of hydraulics — valves control the load, they never hold it",
         "One of nine independent safety barriers",
         "ea", 4, 160, CN),
        ("Proportional flow/pressure valve section per lift",
         "ISO 4413; on manifold M-02",
         None,
         "ea", 4, 40, CN),
        ("Position encoder and synchronisation sensor set per lift",
         "Cross-checked pair per lift",
         None,
         "set", 4, 12, CN),
        ("Pit-edge seal and debris excluder set",
         None, None,
         "set", 4, 40, CN),
        ("Lift control station: hold-to-run, E-stop, key isolation",
         "Hold-to-run is a life-safety requirement",
         None,
         "ea", 4, 32, CN),
        ("Support stands / props, 6 t, for service work",
         "Rated and tagged",
         None,
         "ea", 8, 200, CN),
        ("Design registration dossier for Australian registrable plant",
         "Full design calcs, FMEA, drawings, test reports per AS/NZS 1418.9 for state-regulator design registration",
         "CRITICAL PATH: registration pathway must be resolved before pits are poured",
         "lot", 1, None, CN),
    ]),
    ("M", "Hydraulic power unit & reticulation", [
        ("Hydraulic power unit 15 kW, 100 L/min, 300 L reservoir",
         "ISO 4413; 415 V 3-phase; relief set above 74.9 bar stage-3 working pressure",
         "Full lift of all four cassettes in 117 s at 100 L/min; swept volume 194.8 L",
         "ea", 1, 650, CN),
        ("Proportional manifold block, 4 sections, relief + unloader",
         "Houses L-05 valve sections",
         None,
         "ea", 1, 60, CN),
        ("Pressure transducers and gauges, per stage circuit",
         "4-20 mA, glycerine gauges",
         None,
         "lot", 1, 8, CN),
        ("Pressure and return filtration with clog indicators",
         "10 micron return, 25 micron pressure",
         None,
         "set", 1, 20, CN),
        ("Air-blast oil cooler with thermostat",
         None, None,
         "ea", 1, 35, CN),
        ("Reservoir level / temperature sensor with low-level interlock",
         None, None,
         "set", 1, 4, CN),
        ("Hard piping, zinc-nickel plated steel, DN20/DN12, with fittings",
         "Approx. 90 m total to 4 pits",
         None,
         "lot", 1, 220, CN),
        ("Hose-burst check valve at each cylinder port",
         "ISO 8643 type",
         None,
         "ea", 4, 8, CN),
        ("Flexible hose sets, pit wall to cassette",
         "2-wire braid, tested",
         None,
         "set", 4, 30, CN),
        ("Drain, sampling and bleed points",
         None, None,
         "lot", 1, 5, CN),
        ("First-fill hydraulic oil, ISO VG 46, approx. 300 L",
         "Ship system dry; oil filled locally",
         None,
         "lot", 1, None, LOCAL),
    ]),
    ("N", "Plant sump fit-out", [
        ("Sump steel liner / bund box 2.4 x 1.6 x 1.8 m, for casting-in",
         "6 mm plate, stiffened, seal-welded, bitumen-coated externally",
         "Bund holds 6912 L against a 330 L requirement; the sump IS the bund",
         "ea", 1, 1400, CN),
        ("Trafficable flush cover panels, gasketed, lift-out",
         "Rated for workshop traffic; lifting keys supplied",
         None,
         "set", 1, 380, CN),
        ("Access ladder and retrieval anchor point",
         "AS 1657 ladder; confined-space retrieval point",
         "The sump is a confined space with oil and live electrical gear",
         "set", 1, 35, CN),
        ("Mechanical ventilation: fan 83 m3/h + ducting to rear gable + louvre",
         "12 air changes per hour, runs whenever HPU is energised",
         None,
         "set", 1, 40, CN),
        ("Oil-in-water sensor, sump pump interlock",
         "Interlocked so the sump pump cannot discharge oil",
         "Counted in the 61-point sensor schedule",
         "ea", 1, 2, CN),
        ("Sump pump, automatic, with interlock relay",
         "Submersible, 240 V, oil-interlocked via N-05",
         None,
         "ea", 1, 15, CN),
        ("High-level alarm float and local beacon",
         "Counted in the 61-point sensor schedule",
         None,
         "set", 1, 3, CN),
        ("Sump LED luminaire and IP66 socket outlet",
         "Low-voltage luminaire; RCD-protected outlet",
         None,
         "set", 1, 5, CN),
        ("Confined-space signage and entry-control kit",
         None, None,
         "lot", 1, 3, CN),
    ]),
    ("O", "Floor drainage interfaces", [
        ("Trench drain, galv edge rails + class D grates, 2 runs x 12.2 m",
         "Modular units for casting-in; silt-tight joints",
         "24.4 m total across the workshop bay",
         "m", 24.4, 730, CN),
        ("Trench drain sump outlets with silt baskets",
         "Galv baskets, removable",
         None,
         "ea", 2, 30, CN),
        ("Oil/water triple interceptor, 1000 L",
         "WaterMark-certified product required — local supply",
         "Trade-waste approval by water authority (local)",
         "ea", 1, None, LOCAL),
        ("Interceptor pipework, vent and connection",
         "Local plumbing works",
         None,
         "lot", 1, None, LOCAL),
        ("Stormwater connection pits and first-flush arrangement",
         "Local civil works",
         None,
         "lot", 1, None, LOCAL),
    ]),
    ("P", "Solar array, 52.08 kW", [
        ("PV module 620 W, monocrystalline",
         "IEC 61215 / IEC 61730; must be on the Australian CEC approved-modules list",
         "84 modules = 52.08 kW; 7 rows x 6 columns per slope, landscape; 36.5 kg each, 3062 kg total",
         "ea", 84, 3062, CN),
        ("Roof-mount rail system for PIR sandwich roof",
         "Wind-rated to AS/NZS 1170.2 site category (TBC); approx. 180 m rail",
         "Added dead load with modules 0.1132 kPa",
         "lot", 1, 320, CN),
        ("Panel-interface brackets, sealed fixings through panel to purlins",
         "Panel-manufacturer-approved bracket + seal system",
         None,
         "lot", 1, 90, CN),
        ("Module mid and end clamps",
         "Anodised aluminium",
         None,
         "lot", 1, 25, CN),
        ("DC cable, PV1-F 6 mm2, UV-rated",
         "AS/NZS 5033 compliant installation (by local electrician)",
         "6 strings of 14 modules; string Voc 577 V",
         "m", 560, 45, CN),
        ("MC4 connectors and Y-branch sets",
         "Genuine matched connectors only",
         None,
         "lot", 1, 5, CN),
        ("DC isolators, array and inverter ends",
         "AS 60947.3; IP66",
         None,
         "ea", 4, 8, CN),
        ("Array earthing and bonding kit",
         "Lugs, washers, 6 mm2 earth",
         None,
         "lot", 1, 10, CN),
        ("Roof-to-lean-to cable duct and tray",
         "Galv, UV-stable",
         None,
         "lot", 1, 40, CN),
        ("Array layout, stringing and as-built drawing set",
         "String map 6 x 14; Voc calcs at site minimum temperature",
         None,
         "lot", 1, None, CN),
        ("Spare PV modules, attic stock",
         "Same batch as P-01",
         "Priced option",
         "ea", 2, 73, OPT),
    ]),
    ("Q", "Inverters, battery storage, EV charging", [
        ("String inverter 25 kVA, 3-phase",
         "AS/NZS 4777.2:2020 certified and CEC-listed; grid export limits set at commissioning",
         "2 inverters = 50 kVA; grid connection application by principal (AS/NZS 4777.1)",
         "ea", 2, 90, CN),
        ("Battery cabinet, LFP, outdoor IP55, expandable to 50 kWh each",
         "CEC-listed battery system; installation to AS/NZS 5139 (by local installer)",
         "2 cabinets; building provision to 100 kWh total",
         "ea", 2, 600, CN),
        ("Battery module, LFP, nominal 5 kWh",
         "Same certified system as Q-02",
         "Initial fit 50 kWh (10 modules); expansion to 100 kWh is a future order",
         "ea", 10, 550, CN),
        ("Battery management gateway, energy meter and CT set",
         "Modbus TCP interface to group U controller",
         None,
         "set", 1, 8, CN),
        ("EV charge point 22 kW, Type 2, tethered",
         "AS/NZS 61851.1; RCM marked; OCPP 1.6J",
         "2 points in the lean-to charging bay",
         "ea", 2, 60, CN),
        ("EV load-management interface module",
         "Sheds/derates chargers under the 71.9 kVA site cap",
         None,
         "lot", 1, 3, CN),
        ("Lean-to AC/DC protection and combiner board",
         "AS/NZS 61439 assembly; isolators, breakers, PV and battery protection",
         None,
         "ea", 1, 45, CN),
        ("Surge protection, PV DC and AC sides, T1+T2",
         "AS/NZS 1768 / IEC 61643 devices",
         None,
         "set", 1, 6, CN),
        ("Compliance signage and shutdown procedure set",
         "AS/NZS 4777.2, AS/NZS 5139 and NCC-required labels, engraved",
         None,
         "lot", 1, 2, CN),
    ]),
    ("R", "Plant lean-to, 3.0 x 6.0 m", [
        ("Lean-to frame, SHS 89 x 89 x 5, mono-pitch",
         "AS/NZS 1163 C350, HDG; against right wall",
         "Keeps inverters, batteries and EV charging outside the main volume for fire separation",
         "ea", 1, 700, CN),
        ("Lean-to roof: 100 mm PIR panel + flashings and gutter",
         "Matches group H system",
         None,
         "m2", 20, 250, CN),
        ("Fire-separation wall panel between lean-to and main wall",
         "FRL 60/60/60 wall system, 20 m2; certifier acceptance by principal",
         None,
         "m2", 20, 300, CN),
        ("Equipment mounting: unistrut, plinth frames, cabinet rails",
         "HDG strut + frames for 2 inverters + 2 battery cabinets",
         None,
         "lot", 1, 120, CN),
        ("Mesh enclosure sides with lockable gate",
         "Galv welded mesh, 2 sides + gate",
         None,
         "set", 1, 180, CN),
        ("Protection bollards at charging bay",
         "140 NB concrete-filled (fill local), safety yellow",
         None,
         "ea", 4, 120, CN),
        ("Lean-to downpipe and overflow",
         None, None,
         "set", 1, 15, CN),
    ]),
    ("S", "Main electrical distribution", [
        ("Main switchboard, 415 V 3-phase, 100 A supply, Form 2",
         "AS/NZS 61439; fault level and chassis rating per local supply authority (TBC)",
         "Connected load 84.9 kW against a 71.9 kVA supply — load management mandatory",
         "ea", 1, 280, CN),
        ("Dynamic load-management controller",
         "Sheds EV charging and HPU on priority under the 71.9 kVA cap; Modbus TCP",
         None,
         "ea", 1, 5, CN),
        ("Metering CT chamber and check-metering set",
         "Per local network requirements (pattern TBC)",
         None,
         "set", 1, 15, CN),
        ("Main switchboard surge protection T1+T2",
         "AS/NZS 1768",
         None,
         "set", 1, 4, CN),
        ("Main isolation switch, manual, lockable",
         "Life-safety: manual switchboard isolation independent of any automation",
         None,
         "ea", 1, 6, CN),
        ("Distribution board DB-1, server compartment, with UPS input",
         "AS/NZS 61439",
         None,
         "ea", 1, 40, CN),
        ("Submains and final subcircuit cable, Cu XLPE",
         "AS/NZS 5000.1; approx. 1200 m all sizes",
         "Installation by licensed Australian electrician (local)",
         "lot", 1, 480, CN),
        ("Cable ladder 300 mm and tray 150 mm, HDG, with fittings",
         "Approx. 90 m total",
         None,
         "m", 90, 320, CN),
        ("Conduits, pits and lean-to link ducting",
         "UV-stable rigid conduit + pit set",
         None,
         "lot", 1, 90, CN),
        ("Marshalling terminal strip, single point, fully labelled",
         "Every field device lands on one labelled strip — the control philosophy requires it",
         None,
         "ea", 1, 12, CN),
        ("Earthing and MEN system: electrodes, bonds, test links",
         "AS/NZS 3000 section 5",
         None,
         "lot", 1, 40, CN),
        ("Small power: GPOs, 4 x 32 A 3-phase welding outlets, bench outlet rails",
         "IP-rated where exposed; approx. 28 GPO points + 13 bench rails",
         None,
         "lot", 1, 60, CN),
        ("Under-deck and container-bay reticulation duct",
         "Galv duct, matches B-12 in-bay channels",
         None,
         "lot", 1, 80, CN),
    ]),
    ("T", "Lighting", [
        ("LED highbay 150 W, DALI dimmable",
         "RCM marked; 5000 K; occupancy-linked via group U",
         "Over the 12.192 x 12.192 m bay",
         "ea", 8, 40, CN),
        ("Container-bay LED strip 1.2 m, per bay",
         "RCM marked; DALI",
         "15 bays including server compartment",
         "ea", 15, 30, CN),
        ("Under-deck / perimeter LED strips",
         "RCM marked; DALI",
         None,
         "ea", 12, 24, CN),
        ("External luminaire, sensor-driven, full cut-off",
         "RCM marked; integral or U-group sensor control; dark-sky friendly optics",
         "10 external luminaires per control philosophy — no wall switches",
         "ea", 10, 40, CN),
        ("Emergency and exit luminaires, hardwired",
         "AS/NZS 2293; NOT sensor-controlled (life-safety)",
         None,
         "ea", 12, 24, CN),
        ("Manual lighting override switch at MSB",
         "Life-safety: one manual override independent of the controller",
         None,
         "ea", 1, 1, CN),
        ("DALI gateway and drivers",
         "DALI-2 to Modbus/MQTT bridge",
         None,
         "lot", 1, 6, CN),
        ("Photocell master input",
         "Feeds lux logic; luminaires respond via controller",
         None,
         "ea", 1, 1, CN),
        ("Lighting commissioning: zones, scenes, sensor thresholds",
         "Witnessed test with principal",
         None,
         "lot", 1, None, CN),
    ]),
    ("U", "Sensors, control, security, network", [
        ("Building controller: industrial DIN PC, MQTT broker + Modbus TCP",
         "Open protocols; drops into an existing Home Assistant installation",
         "Runs the no-wall-switch control philosophy; 61 sensor points across 15 types",
         "ea", 1, 3, CN),
        ("Modbus I/O modules, DIN rail",
         "16DI/8DO/8AI mix",
         None,
         "lot", 1, 6, CN),
        ("mmWave presence sensor, internal",
         "24 GHz, RCM",
         "Sensor schedule ref S1",
         "ea", 8, 2, CN),
        ("PIR motion sensor, external, IP65",
         "Sensor schedule ref S2",
         None,
         "ea", 6, 2, CN),
        ("Lux sensor",
         "Sensor schedule ref S3",
         None,
         "ea", 4, 1, CN),
        ("Temperature / humidity sensor",
         "Sensor schedule ref S4",
         None,
         "ea", 6, 1, CN),
        ("CO sensor (vehicle exhaust)",
         "Sensor schedule ref S5; alarm + fan interlock",
         None,
         "ea", 3, 1, CN),
        ("CO2 sensor",
         "Sensor schedule ref S6",
         None,
         "ea", 3, 1, CN),
        ("Smoke / heat detector, interlinked",
         "AS 3786; hardwired alarm path independent of controller",
         "Sensor schedule ref S7",
         "ea", 6, 2, CN),
        ("Door / gate reed contact",
         "Sensor schedule ref S8",
         None,
         "ea", 11, 1, CN),
        ("Roller-door position encoder",
         "Sensor schedule ref S9",
         None,
         "ea", 1, 1, CN),
        ("Water leak probe",
         "Sensor schedule ref S10",
         None,
         "ea", 3, 1, CN),
        ("Energy monitoring CT set, per zone",
         "Sensor schedule ref S13; Modbus meters",
         None,
         "set", 6, 6, CN),
        ("Weather station: wind, rain, external temperature",
         "Sensor schedule ref S14",
         None,
         "ea", 1, 2, CN),
        ("Vibration / tamper sensor, plant areas",
         "Sensor schedule ref S15",
         None,
         "ea", 1, 1, CN),
        ("CCTV camera, external bullet, IP67, IR",
         "ONVIF, PoE",
         "6 external of 9 total cameras",
         "ea", 6, 6, CN),
        ("CCTV camera, internal dome",
         "ONVIF, PoE",
         "3 internal",
         "ea", 3, 2, CN),
        ("NVR / NAS, 8-bay, with camera licences",
         "30-day retention at full resolution",
         None,
         "ea", 1, 8, CN),
        ("Credential reader, OSDP, no keypad",
         "DESFire EV2/EV3; 4 readers — no keypads per brief",
         "Egress is always mechanical (life-safety)",
         "ea", 4, 2, CN),
        ("PoE switch 24-port + patch panel + 1.5 kVA UPS",
         "In server compartment rack",
         None,
         "set", 1, 25, CN),
        ("Structured cabling Cat6A, approx. 900 m, with outlets",
         "Terminated and certified",
         None,
         "lot", 1, 45, CN),
        ("WiFi access point, industrial",
         None, None,
         "ea", 2, 2, CN),
        ("Controller configuration: dashboards, MQTT topics, Home Assistant handover",
         "Config repository handed over; no cloud lock-in",
         None,
         "lot", 1, None, CN),
    ]),
    ("V", "Service reels & workshop services", [
        ("Hose reel: compressed air, 10 mm x 15 m",
         "Spring-retract, swivel mount",
         "Reel bank of five under the front-wall hood",
         "ea", 1, 12, CN),
        ("Cable reel: 240 V 15 A x 18 m",
         "RCM marked, RCD-protected circuit",
         None,
         "ea", 1, 14, CN),
        ("Hose reel: water, 12 mm x 15 m",
         None, None,
         "ea", 1, 12, CN),
        ("Hose reel: oil dispense, 10 mm x 10 m, with metered gun",
         None, None,
         "ea", 1, 15, CN),
        ("Hose reel: grease, 6 mm x 10 m, with gun",
         None, None,
         "ea", 1, 12, CN),
        ("Reel-bank mounting rail and hood interface",
         "Mounts to D-11 hood frame",
         None,
         "set", 1, 25, CN),
        ("Air / water / oil reticulation to reel bank and bays",
         "Press-fit aluminium air line DN25 + copper water + steel oil lines, approx. 70 m",
         None,
         "lot", 1, 120, CN),
        ("Rotary screw compressor 11 kW + dryer + 500 L receiver",
         "Not in the rev C model — priced option, confirm before order",
         "Feeds V-01 and bay drops",
         "set", 1, 420, OPT),
    ]),
    ("W", "Workbenches & bay fit-out", [
        ("Steel workbench 2.2 x 0.75 m, 6 mm top, undershelf",
         "Powder-coated frame, levelling feet",
         "13 workbench bays",
         "ea", 13, 1560, CN),
        ("Bay back-panel tool board with hook kit",
         "Perforated steel, bay width",
         None,
         "ea", 13, 390, CN),
        ("Bench vice 150 mm with mount",
         "Priced option",
         None,
         "ea", 13, 260, OPT),
        ("Server rack 18RU with PDU and cable management",
         "In sealed server compartment",
         None,
         "ea", 1, 45, CN),
        ("Server compartment shelving and storage fit-out",
         None, None,
         "lot", 1, 60, CN),
    ]),
    ("X", "Fixings, sealants, consumables, export packing", [
        ("Chemical anchor system M12/M16 with HDG/SS studs",
         "ETA-assessed system; for hold-downs and plant fixing",
         None,
         "lot", 1, 60, CN),
        ("Masonry anchors and screwbolts, assorted",
         None, None,
         "lot", 1, 40, CN),
        ("General fasteners site allowance: teks, bolts, rivets",
         "Class 4 coated",
         None,
         "lot", 1, 80, CN),
        ("Sealants: PU, neutral-cure silicone, fire-rated penetration sealer",
         None, None,
         "lot", 1, 60, CN),
        ("Touch-up: cold-galvanising and 2-pack repair kits",
         None, None,
         "lot", 1, 30, CN),
        ("EPDM washers, closure strips, foam fillers — spares including 2% attic stock",
         None, None,
         "lot", 1, 25, CN),
        ("Erection mark labels, engraved plant tags, cable tags",
         "Marks match the shop-drawing erection sequence",
         None,
         "lot", 1, 10, CN),
        ("Export packing: cradles, frames, VCI wrap, desiccant, ISPM-15 timber",
         "All timber ISPM-15 stamped; seaworthy packing certificate required",
         "Packed into the three group-A containers plus flat racks as needed",
         "lot", 1, 400, CN),
        ("Lashing and securing of cargo into containers / flat racks",
         "CTU-code compliant lashing plan, photographed",
         None,
         "lot", 1, 120, CN),
    ]),
    ("Y", "Local supply & site works (info only — excluded from this RFQ)", [
        ("Concrete: slab and edge beam (within 90.5 m3 total)",
         "Local supply; engineer's design pending geotech",
         None,
         "lot", 1, None, LOCAL),
        ("Footing pads 1.2 x 1.2 x 0.7 m",
         "12 pads under portal columns",
         None,
         "ea", 12, None, LOCAL),
        ("Lift pit and sump excavation; casting-in of supplied boxes",
         "Uses L-01 pit boxes and N-01 sump liner",
         "Sequenced after lift design registration is resolved",
         "lot", 1, None, LOCAL),
        ("Reinforcement: mesh and bar",
         None, None,
         "lot", 1, None, LOCAL),
        ("Vapour barrier, sawcuts, joint sealing, surface sealer",
         None, None,
         "lot", 1, None, LOCAL),
        ("Holding-down bolt placement and grouting",
         "Uses C-08 cast-in assemblies and templates",
         None,
         "lot", 1, None, LOCAL),
        ("Erection and installation labour, cranage, EWPs",
         "Local erector; supplier site supervision quoted as RFQ option",
         None,
         "lot", 1, None, LOCAL),
        ("Licensed electrical installation and grid connection",
         "AS/NZS 3000; AS/NZS 4777.1 network application",
         None,
         "lot", 1, None, LOCAL),
        ("Plumbing, trade waste and interceptor connection",
         "WaterMark products; water-authority approval",
         None,
         "lot", 1, None, LOCAL),
        ("Certification package: structural, plant registration, building, solar/battery",
         "The 15-item sign-off register in the RFQ brief",
         None,
         "lot", 1, None, LOCAL),
        ("Site data: wind region, terrain category, soil classification",
         "Prerequisite — nothing structural can be certified without it",
         None,
         "lot", 1, None, LOCAL),
        ("Commissioning: crane 125% load test, lift verification, controls witness tests",
         "Supplier commissioning support quoted as RFQ option",
         None,
         "lot", 1, None, LOCAL),
    ]),
]

SENSOR_SCHEDULE = [
    ("S1", "mmWave presence, internal", 8, "U-03"),
    ("S2", "PIR motion, external", 6, "U-04"),
    ("S3", "Lux", 4, "U-05"),
    ("S4", "Temperature / humidity", 6, "U-06"),
    ("S5", "CO (vehicle exhaust)", 3, "U-07"),
    ("S6", "CO2", 3, "U-08"),
    ("S7", "Smoke / heat, interlinked", 6, "U-09"),
    ("S8", "Door / gate contact", 11, "U-10"),
    ("S9", "Roller-door position encoder", 1, "U-11"),
    ("S10", "Water leak probe", 3, "U-12"),
    ("S11", "Oil-in-water (sump interlock)", 1, "N-05"),
    ("S12", "Sump high-level float", 1, "N-07"),
    ("S13", "Energy monitoring CT set", 6, "U-13"),
    ("S14", "Weather station (wind / rain / ext. temp)", 1, "U-14"),
    ("S15", "Vibration / tamper", 1, "U-15"),
]

DIMENSIONS = [
    ("Clear workshop bay", "12.192 x 12.192 m", "containers meet at both rear corners"),
    ("Building envelope", "17.068 x 14.630 m", "footprint 249.7 m2"),
    ("Eave / ridge", "6.896 / 8.432 m", "10.0 deg pitch, equal both slopes"),
    ("Portal frames", "6 at 2.960 m", "17.420 m span, 360 UB 44.7, haunched"),
    ("Envelope", "100 mm PIR", "607 m2 roof, wall and gable"),
    ("Roller door", "4.0 x 4.0 m", "2.875 m clear above the head beam"),
    ("Overhead crane", "2000 kg, 17.420 m span", "hook 5.154 m; reaches 1.914 m over each container top"),
    ("Vehicle lift", "4 x 6 t, 3.0 m stroke", "flush pads on a 4.0 x 3.5 m grid, removable cross beams"),
    ("Hydraulic plant", "15 kW in a floor sump", "2.4 x 1.6 x 1.8 m, bunded, flush covered"),
    ("Container storage deck", "3.001 m level", "max craned item 1.903 m tall"),
    ("Container bays", "15 openings at 2.288 m", "13 workbenches + sealed server compartment"),
    ("Solar array", "52.08 kW", "84 modules, 7 rows x 6 columns per slope"),
    ("Plant lean-to", "3.0 x 6.0 m external", "2 inverters, 2 battery cabinets, 2 x 22 kW EV points"),
    ("Building control", "61 sensor points", "9 cameras, 4 readers, 10 external luminaires, no wall switches"),
    ("Structural steel", "20.7 t", "portals, secondary, crane (groups C-F)"),
    ("Concrete", "90.5 m3", "slab, edge beam, 12 pads, pits, sump — local works"),
    ("Connected load", "84.9 kW", "against a 71.9 kVA supply — load management required"),
]

# ---------------------------------------------------------------- styling
F_BASE = Font(name="Arial", size=10)
F_SMALL = Font(name="Arial", size=9)
F_BOLD = Font(name="Arial", size=10, bold=True)
F_H1 = Font(name="Arial", size=16, bold=True)
F_H2 = Font(name="Arial", size=12, bold=True)
F_WHITE = Font(name="Arial", size=9, bold=True, color="FFFFFF")
F_INPUT = Font(name="Arial", size=10, color="0000FF")
F_GREY = Font(name="Arial", size=10, color="808080")
F_GREY9 = Font(name="Arial", size=9, color="808080")
F_RED = Font(name="Arial", size=10, bold=True, color="A8402A")

FILL_HEAD = PatternFill("solid", fgColor="1F2937")
FILL_GROUP = PatternFill("solid", fgColor="E5E7EB")
FILL_INPUT = PatternFill("solid", fgColor="FFFF00")
FILL_LOCAL = PatternFill("solid", fgColor="F3F4F6")
FILL_OPT = PatternFill("solid", fgColor="FEF3C7")
FILL_TOTAL = PatternFill("solid", fgColor="D1D5DB")

THIN = Side(style="thin", color="C4BDB4")
B_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="top", horizontal="center")
RIGHT = Alignment(vertical="top", horizontal="right")
CENTER = Alignment(vertical="top", horizontal="center")


def style_row(ws, row, cols, font=F_SMALL, border=True):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = font
        if border:
            cell.border = B_ALL
        if cell.alignment.wrap_text is not True:
            cell.alignment = WRAP


def build():
    wb = Workbook()

    # ------------------------------------------------------------ BOM sheet
    ws = wb.active
    ws.title = "BOM"
    headers = ["Item", "Description", "Specification / standard", "Notes", "Unit",
               "Qty", "Nominal mass (kg)", "Scope", "Unit rate (USD)", "Amount (USD)",
               "Tenderer remarks / deviations"]
    widths = [8, 50, 46, 40, 6, 8, 11, 9, 13, 13, 32]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = F_WHITE
        cell.fill = FILL_HEAD
        cell.alignment = WRAP_C
        cell.border = B_ALL
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"

    row = 2
    item_count = 0
    group_rows = {}
    for gcode, gtitle, items in GROUPS:
        ws.cell(row=row, column=1, value=gcode).font = F_BOLD
        gt = ws.cell(row=row, column=2, value=gtitle.upper())
        gt.font = F_BOLD
        for c in range(1, 12):
            ws.cell(row=row, column=c).fill = FILL_GROUP
            ws.cell(row=row, column=c).border = B_ALL
        first = row + 1
        row += 1
        for n, (desc, spec, note, unit, qty, mass, scope) in enumerate(items, start=1):
            code = f"{gcode}-{n:02d}"
            ws.cell(row=row, column=1, value=code)
            ws.cell(row=row, column=2, value=desc)
            ws.cell(row=row, column=3, value=spec or "")
            ws.cell(row=row, column=4, value=note or "")
            ws.cell(row=row, column=5, value=unit)
            qc = ws.cell(row=row, column=6, value=qty)
            qc.number_format = "0.#" if isinstance(qty, float) else "0"
            if mass is not None:
                mc = ws.cell(row=row, column=7, value=mass)
                mc.number_format = "#,##0"
            ws.cell(row=row, column=8, value=scope)
            rate = ws.cell(row=row, column=9)
            amt = ws.cell(row=row, column=10,
                          value=f'=IF(OR($I{row}="",$H{row}="LOCAL"),"",ROUND($F{row}*$I{row},0))')
            amt.number_format = "#,##0"
            ws.cell(row=row, column=11)
            style_row(ws, row, range(1, 12))
            for c in (5, 8):
                ws.cell(row=row, column=c).alignment = CENTER
            for c in (6, 7, 9, 10):
                ws.cell(row=row, column=c).alignment = RIGHT
            if scope == LOCAL:
                for c in range(1, 12):
                    ws.cell(row=row, column=c).font = F_GREY9
                    ws.cell(row=row, column=c).fill = FILL_LOCAL
            else:
                rate.fill = FILL_INPUT
                rate.font = F_INPUT
                rate.number_format = "#,##0.00"
                ws.cell(row=row, column=11).fill = PatternFill("solid", fgColor="FFFDE7")
                if scope == OPT:
                    ws.cell(row=row, column=8).fill = FILL_OPT
            item_count += 1
            row += 1
        group_rows[gcode] = (first, row - 1)

    total_row = row + 1
    ws.cell(row=total_row, column=2, value="TOTAL — CHINA SUPPLY, BASE SCOPE (scope CN)").font = F_BOLD
    tm = ws.cell(row=total_row, column=7, value=f'=SUMIFS($G$2:$G${row - 1},$H$2:$H${row - 1},"CN")')
    tm.number_format = "#,##0"
    ta = ws.cell(row=total_row, column=10,
                 value=f'=SUMIFS($J$2:$J${row - 1},$H$2:$H${row - 1},"CN")')
    ta.number_format = "#,##0"
    opt_row = total_row + 1
    ws.cell(row=opt_row, column=2, value="TOTAL — PRICED OPTIONS (scope CN-OPT)").font = F_BOLD
    oa = ws.cell(row=opt_row, column=10,
                 value=f'=SUMIFS($J$2:$J${row - 1},$H$2:$H${row - 1},"CN-OPT")')
    oa.number_format = "#,##0"
    for r in (total_row, opt_row):
        for c in range(1, 12):
            ws.cell(row=r, column=c).fill = FILL_TOTAL
            ws.cell(row=r, column=c).border = B_ALL
            if ws.cell(row=r, column=c).font.size != 10 or not ws.cell(row=r, column=c).font.bold:
                ws.cell(row=r, column=c).font = F_BOLD
        ws.cell(row=r, column=10).alignment = RIGHT
        ws.cell(row=r, column=7).alignment = RIGHT

    last_data_row = row - 1

    # ------------------------------------------------------- Group summary
    gs = wb.create_sheet("Group summary")
    for i, w in enumerate([8, 52, 8, 12, 15, 15, 15], start=1):
        gs.column_dimensions[get_column_letter(i)].width = w
    gs["A1"] = "GROUP SUMMARY — quote rollup"
    gs["A1"].font = F_H2
    hdr = ["Group", "Title", "Items", "Mass (kg)", "Base CN (USD)", "Options (USD)", "Local (excl.)"]
    for i, h in enumerate(hdr, start=1):
        c = gs.cell(row=3, column=i, value=h)
        c.font = F_WHITE
        c.fill = FILL_HEAD
        c.border = B_ALL
        c.alignment = WRAP_C
    r = 4
    for gcode, gtitle, items in GROUPS:
        gs.cell(row=r, column=1, value=gcode)
        gs.cell(row=r, column=2, value=gtitle)
        gs.cell(row=r, column=3, value=f'=COUNTIFS(BOM!$A$2:$A${last_data_row},"{gcode}-*")')
        m = gs.cell(row=r, column=4, value=f'=SUMIFS(BOM!$G$2:$G${last_data_row},BOM!$A$2:$A${last_data_row},"{gcode}-*")')
        m.number_format = "#,##0"
        b = gs.cell(row=r, column=5, value=(
            f'=SUMIFS(BOM!$J$2:$J${last_data_row},BOM!$A$2:$A${last_data_row},"{gcode}-*",'
            f'BOM!$H$2:$H${last_data_row},"CN")'))
        b.number_format = "#,##0"
        o = gs.cell(row=r, column=6, value=(
            f'=SUMIFS(BOM!$J$2:$J${last_data_row},BOM!$A$2:$A${last_data_row},"{gcode}-*",'
            f'BOM!$H$2:$H${last_data_row},"CN-OPT")'))
        o.number_format = "#,##0"
        loc = gs.cell(row=r, column=7, value=(
            f'=IF(COUNTIFS(BOM!$A$2:$A${last_data_row},"{gcode}-*",'
            f'BOM!$H$2:$H${last_data_row},"LOCAL")>0,"local items","")'))
        style_row(gs, r, range(1, 8), font=F_BASE)
        for c in (3, 4, 5, 6):
            gs.cell(row=r, column=c).alignment = RIGHT
        r += 1
    tot = r
    gs.cell(row=tot, column=2, value="TOTAL").font = F_BOLD
    gs.cell(row=tot, column=3, value=f"=SUM(C4:C{r - 1})").font = F_BOLD
    gs.cell(row=tot, column=4, value=f"=SUM(D4:D{r - 1})").font = F_BOLD
    gs.cell(row=tot, column=4).number_format = "#,##0"
    gs.cell(row=tot, column=5, value=f"=SUM(E4:E{r - 1})").font = F_BOLD
    gs.cell(row=tot, column=5).number_format = "#,##0"
    gs.cell(row=tot, column=6, value=f"=SUM(F4:F{r - 1})").font = F_BOLD
    gs.cell(row=tot, column=6).number_format = "#,##0"
    for c in range(1, 8):
        gs.cell(row=tot, column=c).fill = FILL_TOTAL
        gs.cell(row=tot, column=c).border = B_ALL
    for c in (3, 4, 5, 6):
        gs.cell(row=tot, column=c).alignment = RIGHT

    # commercial block
    cr = tot + 3
    gs.cell(row=cr, column=2, value="COMMERCIAL SUMMARY (tenderer completes yellow cells)").font = F_H2
    lines = [
        ("Subtotal — base scope, ex-works packed", f"=E{tot}", False),
        ("Inland haulage + export formalities, FOB nominated port", None, True),
        ("Sea freight, 3 SOC containers + additional units, to Port of Brisbane", None, True),
        ("Marine insurance (110% CIF value)", None, True),
        ("Third-party pre-shipment inspection allowance", None, True),
        ("TOTAL — CIF Brisbane, base scope", None, "SUM"),
        ("Priced options total (info)", f"=F{tot}", False),
    ]
    start = cr + 1
    for i, (label, val, is_input) in enumerate(lines):
        rr = start + i
        gs.cell(row=rr, column=2, value=label).font = F_BASE
        cell = gs.cell(row=rr, column=5)
        if is_input is True:
            cell.fill = FILL_INPUT
            cell.font = F_INPUT
        elif is_input == "SUM":
            cell.value = f"=SUM(E{start}:E{start + 4})"
            cell.font = F_BOLD
            gs.cell(row=rr, column=2).font = F_BOLD
            for c in range(2, 6):
                gs.cell(row=rr, column=c).fill = FILL_TOTAL
        elif val:
            cell.value = val
        cell.number_format = "#,##0"
        cell.alignment = RIGHT
        for c in range(2, 6):
            gs.cell(row=rr, column=c).border = B_ALL
    note = gs.cell(row=start + len(lines) + 1, column=2,
                   value="Rates are FOB-basis unit prices in USD. State the nominated loading port with your offer. "
                         "DDP options may be offered separately. Nominal masses are for freight planning only — "
                         "final masses come from shop detailing.")
    note.font = F_GREY
    note.alignment = WRAP
    gs.merge_cells(start_row=start + len(lines) + 1, start_column=2,
                   end_row=start + len(lines) + 2, end_column=6)

    # --------------------------------------------------- Sensor schedule
    sn = wb.create_sheet("Sensor schedule")
    for i, w in enumerate([8, 44, 10, 12, 50], start=1):
        sn.column_dimensions[get_column_letter(i)].width = w
    sn["A1"] = "SENSOR & CONTROL POINT SCHEDULE — 61 points, 15 types"
    sn["A1"].font = F_H2
    sn["A2"] = ("Everything runs on sensors — there are no wall switches. Open protocols: MQTT + Modbus TCP, "
                "integrating with an existing Home Assistant installation.")
    sn["A2"].font = F_GREY
    hdr = ["Ref", "Sensor type", "Points", "BOM ref", "Notes"]
    for i, h in enumerate(hdr, start=1):
        c = sn.cell(row=4, column=i, value=h)
        c.font = F_WHITE
        c.fill = FILL_HEAD
        c.border = B_ALL
    r = 5
    for ref, name, qty, bom in SENSOR_SCHEDULE:
        sn.cell(row=r, column=1, value=ref)
        sn.cell(row=r, column=2, value=name)
        sn.cell(row=r, column=3, value=qty).alignment = RIGHT
        sn.cell(row=r, column=4, value=bom)
        style_row(sn, r, range(1, 6), font=F_BASE)
        r += 1
    sn.cell(row=r, column=2, value="TOTAL SENSOR POINTS").font = F_BOLD
    tc = sn.cell(row=r, column=3, value=f"=SUM(C5:C{r - 1})")
    tc.font = F_BOLD
    tc.alignment = RIGHT
    for c in range(1, 6):
        sn.cell(row=r, column=c).fill = FILL_TOTAL
        sn.cell(row=r, column=c).border = B_ALL
    r += 2
    for label, val in [("CCTV cameras", "9  (6 external + 3 internal)"),
                       ("Credential readers (no keypads)", "4"),
                       ("External luminaires, sensor-driven", "10")]:
        sn.cell(row=r, column=2, value=label).font = F_BASE
        sn.cell(row=r, column=3, value=val).font = F_BASE
        r += 1
    r += 1
    ls = sn.cell(row=r, column=2,
                 value="Five functions are deliberately NOT sensor-only (life-safety): emergency lighting, "
                       "switchboard isolation, a manual lighting override, hold-to-run on the crane and lifts, "
                       "and mechanical egress on every door.")
    ls.font = F_RED
    ls.alignment = WRAP
    sn.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=5)

    # ------------------------------------------------------- Dimensions
    dm = wb.create_sheet("Dimensions")
    for i, w in enumerate([34, 26, 58], start=1):
        dm.column_dimensions[get_column_letter(i)].width = w
    dm["A1"] = "SCHEDULE OF PRINCIPAL DIMENSIONS — rev C, 2026-08-24"
    dm["A1"].font = F_H2
    dm["A2"] = ("Every number is computed from one parameter file — the model, this schedule and the BOM "
                "cannot disagree. Status: for engineering. NOT a certified design.")
    dm["A2"].font = F_GREY
    for i, h in enumerate(["Element", "Value", "Note"], start=1):
        c = dm.cell(row=4, column=i, value=h)
        c.font = F_WHITE
        c.fill = FILL_HEAD
        c.border = B_ALL
    r = 5
    for el, val, note in DIMENSIONS:
        dm.cell(row=r, column=1, value=el)
        dm.cell(row=r, column=2, value=val)
        dm.cell(row=r, column=3, value=note)
        style_row(dm, r, range(1, 4), font=F_BASE)
        r += 1
    r += 1
    warn = dm.cell(row=r, column=1,
                   value="Site wind region, terrain category and soil classification are not yet confirmed. "
                         "Wind-rated items (envelope, roller door, PV mounting, structure) are quoted against "
                         "the stated sections; re-pricing after site data lands is handled via the per-tonne and "
                         "per-m2 rates in this workbook.")
    warn.font = F_RED
    warn.alignment = WRAP
    dm.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=3)

    # ------------------------------------------------------------ Cover
    cv = wb.create_sheet("Cover", 0)
    cv.column_dimensions["A"].width = 3
    cv.column_dimensions["B"].width = 30
    cv.column_dimensions["C"].width = 46
    cv.column_dimensions["D"].width = 14
    cv.column_dimensions["E"].width = 14
    cv.column_dimensions["F"].width = 14
    cv["B2"] = "CONTAINER WORKSHOP — BILL OF MATERIALS"
    cv["B2"].font = F_H1
    cv["B3"] = "集装箱车间 — 材料清单 · Request for Quotation 询价书"
    cv["B3"].font = Font(name="Arial", size=12)
    meta = [
        ("RFQ number", RFQ_NO),
        ("Design revision", "C — 2026-08-24, status: for engineering"),
        ("Issued", ISSUE_DATE),
        ("Quotations due", QUOTES_DUE),
        ("Buyer contact", CONTACT),
        ("Delivery basis", "FOB nominated China port; freight to CIF Port of Brisbane, Australia as separate lines"),
        ("Currency", "USD (state exchange assumptions for any CNY components)"),
        ("Validity required", "90 days from submission"),
    ]
    r = 5
    for k, v in meta:
        cv.cell(row=r, column=2, value=k).font = F_BOLD
        cv.cell(row=r, column=3, value=v).font = F_BASE
        r += 1
    r += 1
    cv.cell(row=r, column=2, value="HOW TO COMPLETE THIS WORKBOOK 填写说明").font = F_H2
    r += 1
    instructions = [
        "1. Enter unit rates ONLY in the yellow cells on the BOM sheet (column I) and the freight lines on the Group summary sheet. Amounts and rollups calculate automatically.",
        "2. Quote every line in your package scope. Leave a yellow cell empty only if the line is excluded — and say so in column K.",
        "3. Scope codes: CN = base scope, quote required. CN-OPT = priced option, excluded from the base total. LOCAL = Australian local supply or site works — for information only, do not quote.",
        "4. Record every deviation, substitution or alternative in column K (Tenderer remarks). Substitutions of Australian steel sections (UB/UC to AS/NZS 3679.1) require documented section-property equivalence and are not accepted without written approval.",
        "5. Nominal masses are for freight planning only; final masses come from your shop detailing. Rate and amount columns arrive empty by design: no prices are invented anywhere in this package — they come from supplier quotes.",
        "6. Return this workbook in .xlsx format together with the technical submissions listed in the RFQ brief. The English text governs. 如有歧义，以英文版本为准。",
    ]
    for t in instructions:
        c = cv.cell(row=r, column=2, value=t)
        c.font = F_BASE
        c.alignment = WRAP
        cv.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        cv.row_dimensions[r].height = 30
        r += 1
    r += 1
    cv.cell(row=r, column=2, value="LEGEND & EXAMPLE 图例").font = F_H2
    r += 1
    leg = cv.cell(row=r, column=2, value="Yellow cell = tenderer input")
    leg.fill = FILL_INPUT
    leg.font = F_INPUT
    leg.border = B_ALL
    cv.cell(row=r, column=3, value="Blue text = entered value · Black = calculated · Grey row = LOCAL, not quoted").font = F_BASE
    r += 2
    ex_head = ["Example only", "Unit", "Qty", "Unit rate (USD)", "Amount (USD)"]
    for i, h in enumerate(ex_head):
        c = cv.cell(row=r, column=2 + i, value=h)
        c.font = F_WHITE
        c.fill = FILL_HEAD
        c.border = B_ALL
    r += 1
    cv.cell(row=r, column=2, value="Purlin Z200-19 (illustration of expected format — not a real price)").font = F_BASE
    cv.cell(row=r, column=2).alignment = WRAP
    cv.cell(row=r, column=3, value="m").font = F_BASE
    cv.cell(row=r, column=4, value=100).font = F_BASE
    rate = cv.cell(row=r, column=5, value=8.50)
    rate.fill = FILL_INPUT
    rate.font = F_INPUT
    rate.number_format = "#,##0.00"
    amt = cv.cell(row=r, column=6, value=f"=D{r}*E{r}")
    amt.number_format = "#,##0"
    amt.font = F_BASE
    for c in range(2, 7):
        cv.cell(row=r, column=c).border = B_ALL
    r += 2
    disc = cv.cell(row=r, column=2,
                   value="This BOM derives from the rev C parametric design package (Fusion master: Container Workshop MASTER). "
                         "It is coordinated and dimensionally consistent but NOT a certified design: structural, plant-registration "
                         "and electrical certification happen in Australia against the successful tenderer's shop drawings and data. "
                         "See the RFQ brief for the 15-item sign-off register, quality requirements, packing and shipping terms.")
    disc.font = F_GREY
    disc.alignment = WRAP
    cv.merge_cells(start_row=r, start_column=2, end_row=r + 3, end_column=6)

    # ------------------------------------------------------------ checks
    steel_mass = 0
    cn_mass = 0
    for gcode, _t, items in GROUPS:
        for it in items:
            mass, scope = it[5], it[6]
            if mass:
                if gcode in ("C", "D", "E", "F"):
                    steel_mass += mass
                if scope in (CN, OPT):
                    cn_mass += mass
    assert item_count == 244, f"item count {item_count} != 244"
    assert steel_mass == 20700, f"scheduled steel {steel_mass} != 20700 kg"
    sensor_pts = sum(s[2] for s in SENSOR_SCHEDULE)
    assert sensor_pts == 61, f"sensor points {sensor_pts} != 61"
    assert len(SENSOR_SCHEDULE) == 15
    assert len(GROUPS) == 25

    wb.save("Container-Workshop_BOM_RevC.xlsx")

    summary = {
        "rfq": RFQ_NO, "issued": ISSUE_DATE, "due": QUOTES_DUE,
        "items": item_count, "groups": len(GROUPS),
        "steel_mass_kg": steel_mass, "cn_nominal_mass_kg": cn_mass,
        "group_list": [
            {"code": g, "title": t, "items": len(items),
             "mass": sum(i[5] or 0 for i in items),
             "cn": sum(1 for i in items if i[6] == CN),
             "opt": sum(1 for i in items if i[6] == OPT),
             "local": sum(1 for i in items if i[6] == LOCAL)}
            for g, t, items in GROUPS],
    }
    with open("bom_summary.json", "w") as f:
        json.dump(summary, f, indent=1)
    print(f"OK: {item_count} items, {len(GROUPS)} groups, "
          f"steel C-F {steel_mass / 1000:.1f} t, CN nominal shipped mass {cn_mass / 1000:.1f} t")


if __name__ == "__main__":
    build()
